import os
import frida
import getopt 
import sys
import time
import json
import click
import subprocess
import openpyxl
import datetime
import threading
import random

#  某些机型可能需要关掉selinux
#  adb shell / su / setenforce 0

#*************************************************************************************************************
#   packagename.xlsx   保存当前应用的所有文件操作
#   path.xlsx          保存当前应用的文件删除路径
#   err.txt            报错日志
#*************************************************************************************************************


#*************************************************************************************************************
class Show(object):
    # blue  green   white  red   yellow
    @classmethod
    def error(cls, func, err):
        click.secho('%-20s%-20s%-20s' %('[error]', func, err), fg='red')
        with open("./result/err.txt", "a", encoding='utf-8') as log:
            log.write("********************************************************************************\n")
            log.write(str(datetime.datetime.now()))
            log.write("\n")
            log.write(func)
            log.write("\n")
            log.write(err)
            log.write("\n")

    @classmethod
    def warning(cls, war):
        click.secho('%-20s%-20s' %('[warning]', war), fg='yellow')

    @classmethod
    def info(cls, inf):
        click.secho('%-20s%-20s' %('[info]', inf), fg='white')

    @classmethod
    def user(cls, key, value, change, color):
         click.secho('%-50s%-40s%-20s' %(key, value, change), fg=color)



#*************************************************************************************************************

class Process(object):

    def __init__(self, pkgname):
        self.pid = None
        self.pkgname = pkgname
        self.session = None
        self.script = None
        self.device = frida.get_usb_device(timeout=15)
        pass


    def spawn(self, hook):
        """
        function:  完成应用的重启，附加，JS脚本注入
        """
        try:
            self.pid = self.device.spawn(self.pkgname)
            self.device.resume(self.pid)
            time.sleep(1)
            while True: # 每秒获取一次session 直至成功
                try:
                    self.session = self.device.attach(self.pid)
                except Exception as err:
                    Show.warning('get session ... ...')
                    time.sleep(1)
                    continue
                break
            with open(hook, "r", encoding='utf-8') as f:
                self.script = self.session.create_script(f.read())
            self.script.on("message", on_message) 
            self.script.load()
        except Exception as err:
            Show.error("spawn", str(err))
            exit(2)
        Show.info(self.pkgname + ' spawn completed... ...')



    def attach(self, hook):
        """
        function:  完成应用的附加，JS脚本注入
        """
        try:
            while True: # 每0.5秒获取一次session 直至成功
                try:
                    time.sleep(0.5)
                    self.session = self.device.attach(self.pkgname)
                except:
                    Show.warning('get session ... ...')
                    time.sleep(0.5)
                    continue
                break
            
            with open(hook, "r", encoding='utf-8') as f:
                self.script = self.session.create_script(f.read())
            self.script.on("message", on_message) 
            self.script.load()
        except Exception as err:
            Show.error("attach", str(err) )
            exit(2)
        Show.info(self.pkgname + ' attach completed... ...')


    def unload(self):
        """
        function:  释放script 和 session的资源
        """
        try:
            Show().info('资源释放中... ...')
            self.script.unload()
            self.session.detach()
        except:
            pass


    def get_current_pkg(self):
        """
        function:  获取当前正在运行的应用的包名
        ret：      包名
        """
        try:
            target = self.device.get_frontmost_application() #获取最前端Activity所在进程identifier
            return target.identifier
        except Exception as err:
            Show.error("get_current_pkg", str(err))
            return ""

#*************************************************************************************************************
class CleanData(object):
    _list_file:list = []  # 存放文件操作的 (event + path)
    _list_apis:list = []  # 存放api调用信息 (pai + params)
    _bool_process = False # 是否存在进程清理功能
    _bool_clean = False   # 是否存在垃圾清理功能

    def __init__(self):
        pass


    @classmethod
    def exec_file(cls, data):
        """
        function:  解析并保存文件监控数据
        """
        # print(data)
        event, path = cls.parse_file(data)
        cls._list_file.append({'event':event, 'path':path})


    @classmethod
    def exec_api(cls, data):
        """
        function:  解析并保存api监控数据
        """
        # print(data)
        api, params = cls.parse_api(data)
        cls._list_apis.append({'api':api, 'params':params})


    @classmethod
    def parse_file(cls, data):
        """
        function:  将文件数据拆分
        """
        try:
            return data['event'], data['path']
        except Exception as err:
            Show.error("parse_file", str(err))
            return "event error", "path error"


    @classmethod
    def parse_api(cls, data):
        """
        function:  将api数据拆分
        """
        try:
            return data['api'], data['param']
        except Exception as err:
            Show.error("parse_api", str(err))
            return "api error", "param error"


    @classmethod
    def get_time(cls):
        """
        function:  获取当前时间
        """
        return (str(datetime.datetime.now())).split('.')[0]


    @classmethod
    def save(cls, pkgname):
        """
        function:  将结果保存到xlsx文件
        """
        xlsx = "./result/" + pkgname + ".xlsx"
        try: # 保存应用数据
            excel_result = openpyxl.load_workbook(xlsx)
            sheel_file = excel_result["file"]
            sheel_api = excel_result["api"]

            rows = sheel_file.max_row + 1
            for file in cls._list_file:
                sheel_file.cell(row = rows, column = 1, value = cls.get_time())
                sheel_file.cell(row = rows, column = 2, value = file['event'])
                sheel_file.cell(row = rows, column = 3, value = file['path'])               
                rows += 1

            rows = sheel_api.max_row + 1
            for apis in cls._list_apis:
                sheel_api.cell(row = rows, column = 1, value = cls.get_time())
                sheel_api.cell(row = rows, column = 2, value = apis['api'])
                sheel_api.cell(row = rows, column = 3, value = apis['params'])               
                rows += 1

            excel_result.save(xlsx)
        except Exception as err:
            Show.error("save_file", str(err))

        try: # 保存路径到运营表
            excel_path = openpyxl.load_workbook('./result/path.xlsx')
            sheel_path = excel_path["file"]
            rows = sheel_path.max_row + 1
            for file in cls._list_file:
                #print('event ----> ')
                #print(file)
                sheel_path.cell(row = rows, column = 1, value = pkgname)
                sheel_path.cell(row = rows, column = 2, value = file['event'])
                sheel_path.cell(row = rows, column = 3, value = file['path'])  
                rows += 1
            excel_path.save('./result/path.xlsx')
        except Exception as err:
            Show.error("save_file", str(err))


    @classmethod
    def get_file_data(cls):
        return cls._list_file


    @classmethod
    def get_api_data(cls):
        return cls._list_apis


    @classmethod
    def parse_data(cls, pkgname):
        """
        function:  解析数据，打印结果
        """
        cls.parse_api_data()
        cls.parse_file_data(pkgname)
        #if cls._bool_process:
        #    Show().info('该应用具备进程清理功能')
        #if cls._bool_clean:
        #    Show().info('该应用具备垃圾清理功能')
        #pass


    @classmethod 
    def is_cache(cls, path, pkgname):
        """
        function:  判断当前操作的文件是否为清理应用自己目录下的文件
        """
        if pkgname == '':
            return 1

        if (path.find(pkgname) != -1):
            return 0
        else:
            return 1


    @classmethod
    def parse_file_data(cls, pkgname):
        """
        function:  解析文件监控数据，并打印
        """
        total = 0
        Show().info('监控到的文件访问事件为如下：')
        Show().user(key='Event', value="Path", change="", color="yellow")
        for files in cls._list_file:
            Show().user(key=files['event'], value=files['path'], change="", color="green")
            if (files['event'].find('DELETE') != -1):
                total += cls.is_cache(files['path'], pkgname)
        if (total >= 40):
            cls._bool_clean = True
        pass


    @classmethod
    def parse_api_data(cls): # 目前只监控进程清理的api，测试发现hook其他api会导致应用清理功能出现异常，甚至导致应用闪退
        """
        function:  解析api调用数据，并打印
        """
        total = 0
        Show().info('监控到的API调用情况为如下：')
        Show().user(key='Api', value="Param", change='Function', color="yellow")
        for apis in cls._list_apis:
            if (apis['api'].find('getPackageSizeInfo') != -1):
                Show().user(key='获取应用缓存', value=apis['api'], change=apis['params'], color="green")
            elif (apis['api'].find('getMemoryInfo') != -1):
                Show().user(key='获取系统内存', value=apis['api'], change=apis['params'], color="green")
            elif (apis['api'].find('getProcessMemoryInfo') != -1):
                Show().user(key='获取进程内存', value=apis['api'], change=apis['params'], color="green")
            elif (apis['api'].find('killBackgroundProcesses') != -1):
                total += 1
                Show().user(key=apis['api'], value=apis['params'], change='结束进程' ,color="green")
            elif (apis['api'].find('getRunningAppProcesses') != -1):
                Show().user(key=apis['api'], value=apis['params'], change='获取进程列表', color="green")
            elif (apis['api'].find('ContentResolver.delete') != -1):
                Show().user(key='删除文件', value=apis['api'], change=apis['params'], color="green")
            elif (apis['api'].find('File.delete') != -1):
                Show().user(key='删除文件', value=apis['api'], change=apis['params'], color="green")
        if (total >= 5):
            cls._bool_process = True
        pass


    @classmethod
    def save_file(cls, event:str, path:str): # 废弃
        """
        function:  保存文件监控数据
        """
        try:
            excel_result = openpyxl.load_workbook("./result/result.xlsx")
            sheel_file = excel_result["file"]
            rows = sheel_file.max_row + 1
            sheel_file.cell(row = rows, column = 1, value = cls.get_time())
            sheel_file.cell(row = rows, column = 2, value = event)
            sheel_file.cell(row = rows, column = 3, value = path)
            excel_result.save("./result/result.xlsx")
        except Exception as err:
            excel_result.save("./result/result.xlsx")
            Show.error("save_file", str(err))


    @classmethod
    def save_apis(cls, event:str, path:str): # 废弃
        """
        function:  保存api调用数据
        """
        try:
            excel_result = openpyxl.load_workbook("./result/result.xlsx")
            sheel_api = excel_result["api"]
            rows = sheel_api.max_row + 1
            sheel_api.cell(row = rows, column = 1, value = cls.get_time())
            sheel_api.cell(row = rows, column = 2, value = event)
            sheel_api.cell(row = rows, column = 3, value = path)
            excel_result.save("./result/result.xlsx")
        except Exception as err:
            excel_result.save("./result/result.xlsx")
            Show.error("save_file", str(err))


#*************************************************************************************************************
class WorkThread(threading.Thread):
    _begin = False
    _pkgname = ""
    _list_process = []
    _memory_start = 0
    _memory_end = 0
    _disk_start = 0
    _disk_end = 0
    _process_start = 0
    _process_end = 0

    def __init__(self):
        threading.Thread.__init__(self)
        pass

    def run(self):
        Show.info('获取初始数据 ... ...')
        # 获取初始环境
        self.get_start_performance()
        # 打开开关
        self.set_begin()
        Show.info('监控中 ... ...')
        time.sleep(20)

        while True:
            if self.get_begin():
                if self.choose_func('是否结束并统计数据？ [y / n]'):
                    # 获取清理后环境
                    self.get_end_performance()
                    # 释放资源
                    self.release_process()
                    # 解析数据
                    CleanData().parse_data(self._pkgname)
                    # 显示性能对比
                    self.show_performance_change()
                    # 清理结束 保存数据
                    CleanData().save(self._pkgname)
                    return
                else:
                    continue
        pass


    
    def show_performance_change(self):
        '''
        function:  展示清理前后性能对比
        '''
        Show().info('清理前后性能对比')
        Show().user(key='Target', value='Start', change='End', color="yellow")
        Show().user(key='MemoryAvailable', value=str(self._memory_start), change=str(self._memory_end), color="green")
        Show().user(key='DiskAvailable', value=str(self._disk_start), change=str(self._disk_end), color="green")
        Show().user(key='Process', value=str(self._process_start), change=str(self._process_end), color="green")


    def deal_process(self, process:list):
        '''
        function:  统计进程个数
        '''
        return len(process) - 2 # 去掉标题行和空行


    def format_memory(self, info):
        '''
        function:  获取内存数据
        '''
        value = info.split(': ')[1].strip()
        value = value.split('\\')[0]
        value = value.split(' ')[0]
        return int(value)


    def deal_memory(self, mem:list):
        '''
        function:  处理内存信息，筛选出可用内存
        '''
        for value in mem:
            if (str(value).find('MemAvailable') != -1):
                return self.format_memory(str(value))
        return 0


    def deal_disk(self, disk:list):
        '''
        function:  处理磁盘信息，筛选出/storage/emulated大小
        '''
        for value in disk:
            if (str(value).find('emulated') != -1):
                if (str(value).find('emulated') != -1):
                    return self.format_disk(str(value))
        return 0


    def format_disk(self, info):
        '''
        function:  获取磁盘数据
        '''
        values = info.split(' ')
        count = 0
        for value in values:
            if (value != ""):
                count += 1
            if count == 4:
                return int(value)
        return 0


    def get_start_performance(self):
        '''
        function:  获取清理前手机性能
        '''
        self._memory_start = self.deal_memory(Shell().get_shell('adb shell cat /proc/meminfo'))
        self._disk_start = self.deal_disk(Shell().get_shell('adb shell df -a'))
        self._process_start = self.deal_process(Shell().get_shell('adb shell ps -A'))
        pass


    def get_end_performance(self):
        '''
        function:  获取清理后手机性能
        '''
        self._memory_end = self.deal_memory(Shell().get_shell('adb shell cat /proc/meminfo'))
        self._disk_end = self.deal_disk(Shell().get_shell('adb shell df -a'))
        self._process_end = self.deal_process(Shell().get_shell('adb shell ps -A'))
        pass


    @classmethod
    def set_process(cls, processes:list):
        cls._list_process = processes


    @classmethod
    def release_process(cls):
        """
        function:  释放被hook进程的资源
        """
        for pro in cls._list_process:
            pro.unload()


    @classmethod
    def set_pkgname(cls, pkg):
        cls._pkgname = pkg


    @classmethod
    def set_begin(cls):
        cls._begin = True


    @classmethod
    def get_begin(cls) -> bool:
        return cls._begin


    def choose_func(self, hint):
        """
        function:  获取用户输入
        """
        choose = input(hint)
        sys.stdout.flush()
        if  choose == 'n':
            return False
        elif choose == 'y':
            return True



    def parse_xlsx(self) -> list: #此函数未使用
        """
        function:  解析xlsx
        """
        try:
            excel_activity = openpyxl.load_workbook("activity.xlsx")
            sheel_activity = excel_activity["sheel1"]
            num_row_total = sheel_activity.max_row  # 获取总行数
            list_hash = []
            for rows in range(2, num_row_total + 1):
                hash = sheel_activity.cell(row = rows, column = 1).value
                pkgname = sheel_activity.cell(row = rows, column = 2).value
                activity = sheel_activity.cell(row = rows, column = 3).value
                list_hash.append({"hash":hash, "pkg":pkgname, "activity":activity})
            excel_activity.save("activity.xlsx")
            return list_hash
        except Exception as err:
            Show.error("parse_xlsx", str(err))
            sys.exit(2)
        else:
            Show.info("xlsx 文件解析完成")

#*************************************************************************************************************
class Configer(object):
    def __init__(self):
        #if not self.has_server():
        #    self.push_server()
        #    self.chmod_server()
        #self.forward_frida()
        #self.start_server_popen()
        self.config()
        Show.info('初始化环境配置 ... ...') # 要改成检测手机环境 检测手机编号  json文件中写入编号  每次查找是否有此编号{"device":["455666", "fs52564564"]}
        self.check_config()
        pass

    def get_device(self):
        device = None
        cmd = r'adb devices'         # % apk_file
        pr = subprocess.Popen(cmd, stdout=subprocess.PIPE, shell=True)
        pr.wait()                    # 不会马上返回输出的命令，需要等待
        out = pr.stdout.readlines()  # out = pr.stdout.read().decode("UTF-8")
        for i in (out)[1:-1]:
            device = str(i).split("\\")[0].split("'")[-1]
            Show.info('已连接设备 -> ' + device)
            if (i == 2):
                Show.error('get_device', '当前工具只支持单个设备，请拔出多余设备！')
                exit(2)
        return device


    def popup_AutoStartManagementActivity(self):
        try:
            model = str(Shell.get_shell('adb shell getprop ro.product.model')[0])
            if ((model.find('Redmi') != -1) or (model.find('xiaomi') != -1) or (model.find('Xiaomi') != -1)):
                Shell.excute_shell_strong(['su\n', 'am start -n com.miui.securitycenter/com.miui.permcenter.autostart.AutoStartManagementActivity\n'])
            elif (model.find('AOSP') != -1):
                pass
            elif (model.find('Pixel') != -1):
                pass
        except Exception as err:
            Show.error('popup_AutoStartManagementActivity', str(err))


    def start_process(self):
        """
        function:  主动开启垃圾进程
        """
        for i in range(1, 16):
            Shell.excute_shell_strong(['su\n', 'am start -n com.test' + str(i) + '/com.test' + str(i) + '.MainActivity' + str(i) + '\n'])
            time.sleep(1)
            Shell.back_desk()
        Shell.excute_shell_strong(['su\n', 'am start -n com.filemonitor/com.filemonitor.MainActivity\n'])


    def check_config(self):
        """
        function:  检查配置
        """
        data = None
        device = None
        try:
            device = self.get_device()
            with open("init/config.json", "r", encoding='utf-8') as f:
                data = json.load(f)
            if device not in data['device']:
                # 安装apk
                with open("install.bat", "w+", encoding='utf-8') as bat:
                    bat.write('for %%i in (./init/apk/*.apk) do adb install ./init/apk/%%i')
                    bat.write('\n')
                    bat.write('exit')
                    bat.write('\n')
                    time.sleep(0.1)
                os.system('start install.bat')
                time.sleep(30)
                os.remove('install.bat')

                # 启动垃圾进程
                self.start_process()

                # push server
                with open("push.bat", "w+", encoding='utf-8') as push:
                    push.write('echo "push server ... ..."')
                    push.write('\n')
                    push.write('adb push ./init/server-14-1-3 /data/local/tmp') 
                    push.write('\n')
                    push.write('echo "push over ... ..."')
                    push.write('\n')
                    push.write('exit')
                    push.write('\n')  
                    time.sleep(0.1)
                os.system('start push.bat')
                time.sleep(3)
                os.remove('push.bat')

                # chmod 777
                with open("chmod.bat", "w+", encoding='utf-8') as chmod:
                    chmod.write('echo "chmod 777 server ... ..."')
                    chmod.write('\n')
                    chmod.write('echo su > 123.txt')
                    chmod.write('\n')
                    chmod.write('echo chmod 777 /data/local/tmp/server-14-1-3 >> 123.txt')
                    chmod.write('\n')    
                    chmod.write('adb shell < 123.txt')
                    chmod.write('\n')  
                    chmod.write('del 123.txt')
                    chmod.write('\n')  
                    chmod.write('echo "chmod over"')
                    chmod.write('\n')  
                    chmod.write('exit')
                    chmod.write('\n') 
                os.system('start chmod.bat')
                time.sleep(1)
                os.remove('chmod.bat')

                # 更新配置文件
                data['device'].append(device)
                with open("init/config.json", "w", encoding='utf-8') as f:
                    json.dump(data, f)

                # 打开server
                with open("server.bat", "w", encoding='utf-8') as push:
                    push.write('echo "server server ... ..."')
                    push.write('\n')
                    push.write('adb forward tcp:27042 tcp:27042')
                    push.write('\n')
                    push.write('adb shell < ./init/shell.txt')
                    push.write('\n')    
                with open("init/shell.txt", "w", encoding='utf-8') as shell:
                    shell.write('su')
                    shell.write('\n')
                    shell.write('./data/local/tmp/server-14-1-3')
                    shell.write('\n')
                os.system('start server.bat')
                time.sleep(5)

                # 弹出自启动授权页
                Show.info('初次安装垃圾应用 [Application]，为达到保活效果，请授予[Application 1-15] 自启动权限')
                self.popup_AutoStartManagementActivity()
            else:
                # 手机已经初始化
                self.start_server()
        except Exception as err:
            Show.error('check_config', str(err))
            exit(2)


    def start_server(self):
        '''
        function:  启动server
        '''
        if not self.check_server():
            os.system('start server.bat')
        else:
            Show.info('检查到服务器窗口已开启...')


    def check_server(self):
        '''
        function:  检查server是否启动
        '''
        try:
            list_ps = Shell.get_shell('adb shell "ps | grep server-14-1-3"')
            for ps in list_ps:
                if (str(ps)).find('server-14-1-3') != -1:
                    return True
        except Exception as err:
            Show.error('check_server', str(err))
        return False


  
    def config(self):
        '''
        function:  检查文件
        '''
        # 检查file文件夹是否存在
        if not os.path.exists("./result"):
            os.mkdir("./result")
        if not os.path.exists("./result/path.xlsx"):
            workbook = openpyxl.Workbook()
            file = workbook.create_sheet("file", 0)
            file.cell(1, 1, "Packagename")
            file.cell(1, 2, "Event")
            file.cell(1, 3, "Path")
            workbook.save("./result/path.xlsx")
#*************************************************************************************************************

class Shell(object):
    def __init__(self):
        pass


    @classmethod
    def get_shell(cls, cmd:str) -> list:
        '''
        function:  执行adb单条命令，并获取命令行返回值
        '''
        try:
            obj = subprocess.Popen(cmd, shell = True, stdin=subprocess.PIPE, stdout=subprocess.PIPE ,stderr=subprocess.PIPE)
        except Exception as err:
            Show.error('get_shell', str(err))
            return None
        return obj.stdout.readlines()


    @classmethod
    def get_shell_strong(cls, cmd:list) -> list:
        '''
        function:  执行adb多条命令，并获取命令行返回值
        '''
        info = None
        try:
            obj = subprocess.Popen(['adb', 'shell'], shell = True, stdin=subprocess.PIPE, stdout=subprocess.PIPE ,stderr=subprocess.PIPE)
            for line in cmd:
                obj.stdin.write(line.encode('utf-8'))
            info,err = obj.communicate()
        except Exception as err:
            Show.error('get_shell_strong', str(err))
            return info
        return (str(info.decode('gbk'))).split('\n')


    @classmethod
    def excute_shell(cls, cmd:str):
        '''
        function:  执行adb单条命令
        '''
        try:
            obj = subprocess.Popen(cmd, shell = True, stdin=subprocess.PIPE, stdout=subprocess.PIPE ,stderr=subprocess.PIPE)
            info,err = obj.communicate() # 必须要有这行读取结果 不然会出错
        except Exception as err:
            Show.error('excute_shell', str(err))



    @classmethod
    def excute_shell_strong(cls, cmd:list):
        '''
        function:  执行adb多条命令
        '''
        try:
            obj = subprocess.Popen(['adb', 'shell'], shell = True, stdin=subprocess.PIPE, stdout=subprocess.PIPE ,stderr=subprocess.PIPE)
            for line in cmd:
                obj.stdin.write(line.encode('utf-8'))
            info,err = obj.communicate() # 必须要有这行读取结果 不然会出错
        except Exception as err:
            Show.error('excute_shell_strong', str(err))


    @classmethod
    def back_desk(cls):
        '''
        function:  返回桌面并右滑
        '''
        try:
            time.sleep(0.5)
            os.system("adb shell input keyevent 3") # 模拟按下home键
            time.sleep(0.1)
            os.system("adb shell input keyevent 3") # 模拟按下home键
            time.sleep(0.1)
            os.system("adb shell input keyevent 3") # 模拟按下home键
            time.sleep(0.1)
            os.system("adb shell input swipe 500 600 100 600") # 右滑回到原桌面页
        except Exception as err:
            Show.error("back_desk", str(err))

#*************************************************************************************************************
def on_message(message, data):
    """
    function:  消息回调，用于接收并处理JS中发送的数据信息
    """
    if message['type'] == 'send':
        if WorkThread().get_begin():
            if message['payload'][:7] == 'apis:::':
                #Show.warning(' apis ')
                jsonpacket = json.loads(message['payload'][7:])
                CleanData().exec_api(jsonpacket)
            elif message['payload'][:7] == 'file:::':
                #Show.warning(' file ')
                jsonpacket = json.loads(message['payload'][7:])
                CleanData().exec_file(jsonpacket)
    else:
        #print("*****[frida hook]***** : " + str(message))
        pass


#*************************************************************************************************************
banner1 = [ 
    '**************************************************************************************************************'
    ]

banner = [ 
    '      *************    ***              *************         ***         ***        ***     ',
    '     **************    ***              *************      ***   ***      *****      ***     ',
    '     ***               ***              ***              ***       ***    ***  **    ***     ',
    '     ***               ***              *************    *************    ***   **   ***     ',
    '     ***               ***              *************    *************    ***    **  ***     ',
    '     ***               ***              ***              ***       ***    ***     ** ***     ',
    '     **************    *************    *************    ***       ***    ***      *****     ',
    '      *************     ************    *************    ***       ***    ***        ***     ',
    ]
#*************************************************************************************************************
class Entry(object):
    def __init__(self, argv:list):
        self.pkgname = None
        self.device = None
        self.parseThread = None
        self.model = False
        self.argv(argv)
        pass



    def show_banner(self):
        """
        function:  显示banner
        """
        colors = ['red', 'green', 'blue', 'cyan', 'magenta', "yellow", "white"]
        try:
            columns = os.get_terminal_size().columns
            if columns >= len(banner[1]):
                for line in banner:
                    if line:
                        fill = int((columns - len(line)) / 2) - 1
                        line = line[0] * fill + line
                        line += line[-1] * fill
                    click.secho(line, fg=random.choice(colors))
        except:
            pass
        time.sleep(2)
        pass


    def show_help(self):
        """
        function:  帮助文档
        """
        tips = [
            '*******************************************************************************',
            '  python main.py -[s/a] -p [packagename]',
            '  -p 后面跟要监控的包名 '
            '  -a attach模式 需预先打开FileMonitor 和 待测试应用'
            '  -s spawn 模式 脚本会自动打开Filemonitor和 待测试应用'
            '*******************************************************************************'
        ]
        for line in tips:
            click.secho(line, fg='yellow')



    def argv(self, argv):
        """
        function:  解析命令行参数
        """
        try:
            opts, args = getopt.getopt(argv, "hsap:")
            for arg, value in opts:
                if arg == '-h':
                    self.show_help()
                    exit(0)
                elif arg == '-p':
                    self.pkgname = value
                elif arg == '-a':
                    self.model = False
                elif arg == '-s':
                    self.model = True
        except getopt.GetoptError:
            self.show_help()
            exit(2)


    def send_cache(self):
        """
        function:  主动投递缓存文件
        """
        for i in range(1, 30):
            Shell().excute_shell('adb push ./init/cache /storage/emulated/0/cachelog/cache' + str(i) + '.log')
            Shell().excute_shell('adb push ./init/cache /storage/emulated/0/ByteDownload/app' + str(i) + '.apk')


    def start_process(self):
        """
        function:  主动开启垃圾进程
        """
        for i in range(1, 16):
            Shell.excute_shell_strong(['su\n', 'am start -n com.test' + str(i) + '/com.test' + str(i) + '.MainActivity' + str(i) + '\n'])
            time.sleep(1)
            Shell.back_desk()



    def get_activity(self):
        '''
        function:  获取当前活动页的pkg  activity
        '''
        try:
            value = str(Shell().get_shell_strong(['dumpsys activity | grep "mResume"\n', 'exit\n'])[0])
        except Exception as err:
            Show.error('get_activity', str(err))
            Show.error('get_activity', '请检查手机是否关屏 / adb是否能正常连接')
            exit(2)
        else:
            return value.split('u0 ')[1].split(' ')[0].split('/')  


    def get_hook(self):
        """
        function:  获取监控应用
        """
        self.pkgname = self.choose_pkg()


    def choose_pkg(self):
        """
        function:  选择监控应用
        """
        while True:
            pkg, activity = self.get_activity()
            hint = '是否开始监控当前窗口  packagename=[ ' + pkg + ' ]    Activity=[ ' + activity + ' ] ? [y/n]' 
            choose = input(hint)
            sys.stdout.flush()
            if  choose == 'n':
                continue
            elif choose == 'y':
                return pkg


    def create_xlsx(self, pkgname):
        # 检查并创建pkgname.xlsx
        if not os.path.exists("./result/" + pkgname + ".xlsx"):
            workbook = openpyxl.Workbook()
            file = workbook.create_sheet("file", 0)
            file.cell(1, 1, "Time")
            file.cell(1, 2, "Event")
            file.cell(1, 3, "Path")
            api = workbook.create_sheet("api", 0)
            api.cell(1, 1, "Time")
            api.cell(1, 2, "API")
            api.cell(1, 3, "Param")
            workbook.save("./result/" + pkgname + ".xlsx")


    def excute(self):
        """
        function:  主功能函数
        """
        # 展示banner
        self.show_banner()

        # 初始化
        Configer()

        # 投递垃圾文件
        Show.info('正在投递垃圾文件 ... ...')
        self.send_cache()

        # 垃圾进程无需再手动开启，15个垃圾应用写成保活状态，能自启
        # Show.info('正在开启垃圾进程 ... ...')
        # self.start_process()

        # 获取当前顶层应用
        self.get_hook()

        # 创建结果记录文件
        self.create_xlsx(self.pkgname)

        # hook 文件监控器
        Show.info('Hook FileMonitor ... ...')
        process_filemonitor = Process("com.filemonitor")
        if self.model:
            process_filemonitor.spawn("./init/filemonitor.js")
        else:
            process_filemonitor.attach("./init/filemonitor.js")

        # 返回桌面
        if self.model:
            Shell().back_desk()

        # hook 目标应用
        Show.info('Hook ' + self.pkgname + ' ... ...')
        process_cleanapp = Process(self.pkgname)
        if self.model:
            process_cleanapp.spawn("./init/cleanapp.js")
        else:
            process_cleanapp.attach("./init/cleanapp.js")

        Show.info('Initialization completed.. ...')

        # 开启交互线程
        self.workThread = WorkThread()
        self.workThread.set_pkgname(self.pkgname)
        self.workThread.set_process([process_filemonitor, process_cleanapp]) # 保存对象用于释放资源
        self.workThread.start()


#*************************************************************************************************************
if __name__ == "__main__":
    entry = Entry(sys.argv[1:])
    entry.excute()

    

